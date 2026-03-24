from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "docs" / "US Army_all_states_in_one.xlsx"
OUT_DIR = ROOT / "scripts" / "seed_data"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def main() -> None:
    wb = load_workbook(INPUT_XLSX)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        rec = {}
        for i, value in enumerate(row):
            key = headers[i] if i < len(headers) and headers[i] else f"col_{i}"
            rec[key] = value
        records.append(rec)
    with open(OUT_DIR / "raw_excel_rows.json", "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, default=str)
    print(f"Exported {len(records)} rows to {OUT_DIR / 'raw_excel_rows.json'}")


if __name__ == "__main__":
    main()
